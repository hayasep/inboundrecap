import os
import io
import json
import math
import smtplib
from email.message import EmailMessage
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell.cell import MergedCell
from dotenv import load_dotenv

# Load env (Render secret file first, then local .env)
for p in (Path("/etc/secrets/.env"), Path(__file__).with_name(".env")):
    if p.exists():
        load_dotenv(p, override=True)

# ---------------- Config ----------------
TEMPLATE_PATH = os.environ.get("EXCEL_TEMPLATE_PATH", "backstock report.xlsx")
TEMPLATE_PATH = str(Path(TEMPLATE_PATH).resolve())
SHEET_NAME = os.environ.get("SHEET_NAME", "")

SMTP_SERVER   = os.environ.get("SMTP_SERVER", "")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SENDER_EMAIL  = os.environ.get("SENDER_EMAIL", "")
DEFAULT_RECIPIENT = os.environ.get("DEFAULT_RECIPIENT", "")

ALLOW_BLANK_OVERWRITE = os.environ.get("ALLOW_BLANK_OVERWRITE", "0").lower() in ("1", "true", "yes")
UNIFORM_ROW_HEIGHT_PX = os.environ.get("UNIFORM_ROW_HEIGHT_PX")
DEFAULT_COL_WIDTH_PX  = os.environ.get("DEFAULT_COL_WIDTH_PX")

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "change-me")

# ---------------- Helpers ----------------
def _pt_to_px(pt: float) -> int:
    return int(round(float(pt) * (96.0 / 72.0)))

def _px_to_pt(px: float) -> float:
    return float(px) * 72.0 / 96.0

def _col_char_width(ws, col_index: int) -> float:
    letter = get_column_letter(col_index)
    cd = ws.column_dimensions.get(letter)
    if cd and cd.width:
        return float(cd.width)
    return float(getattr(ws.sheet_format, "defaultColWidth", 8.43))

def _estimate_needed_lines(text: str, chars_per_line: int) -> int:
    if chars_per_line <= 0:
        chars_per_line = 1
    parts = str(text).splitlines() or [""]
    lines = 0
    for part in parts:
        lines += max(1, math.ceil(len(part) / chars_per_line))
    return max(1, lines)

def _last_used_bounds(ws):
    last_r, last_c = 0, 0
    for r in range(1, ws.max_row + 1):
        row_has = False
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                row_has = True
                last_c = max(last_c, c)
        if row_has:
            last_r = r
    return last_r, max(last_c, 1)

def _uniform_row_height_pt(ws):
    if UNIFORM_ROW_HEIGHT_PX:
        return _px_to_pt(float(UNIFORM_ROW_HEIGHT_PX))
    rd1 = ws.row_dimensions.get(1)
    if rd1 and rd1.height:
        return float(rd1.height)
    drh = getattr(ws.sheet_format, "defaultRowHeight", None)
    if drh:
        return float(drh)
    return 15.0

# ---------- Merge utilities ----------
def snapshot_merges(ws):
    return [(mr.min_row, mr.min_col, mr.max_row, mr.max_col) for mr in ws.merged_cells.ranges]

def unmerge_all(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(range_string=mr.coord)

def reapply_merges(ws, merges):
    for r1, c1, r2, c2 in merges:
        if r1 <= r2 and c1 <= c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def adjust_merges_row_offset(merges, row_offset):
    if not row_offset:
        return merges
    return [(r1 + row_offset, c1, r2 + row_offset, c2) for (r1, c1, r2, c2) in merges]

def map_to_anchor_with_snapshot(r, c, merges_snapshot):
    for r1, c1, r2, c2 in merges_snapshot:
        if r1 <= r <= r2 and c1 <= c <= c2:
            return r1, c1
    return r, c

# --------- Web grid JSON (Luckysheet-style) ---------
def wb_to_luckysheet_json(xlsx_path, sheet_name=""):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    if UNIFORM_ROW_HEIGHT_PX:
        uniform_row_px = int(UNIFORM_ROW_HEIGHT_PX)
    else:
        first_rd = ws.row_dimensions.get(1)
        if first_rd and first_rd.height:
            uniform_row_px = _pt_to_px(first_rd.height)
        else:
            drh_pt = getattr(ws.sheet_format, "defaultRowHeight", None)
            uniform_row_px = _pt_to_px(drh_pt) if drh_pt else 20

    default_col_px = None
    dcw_chars = getattr(ws.sheet_format, "defaultColWidth", None
                        )
    if dcw_chars:
        default_col_px = int(round(float(dcw_chars) * 7))
    if DEFAULT_COL_WIDTH_PX:
        default_col_px = int(DEFAULT_COL_WIDTH_PX)

    data = []
    for r in range(1, max_row + 1):
        row_arr = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_arr.append({"v": v} if v not in (None, "") else None)
        data.append(row_arr)

    merges = {}
    for mr in ws.merged_cells.ranges:
        r0 = mr.min_row - 1
        c0 = mr.min_col - 1
        rs = mr.max_row - mr.min_row + 1
        cs = mr.max_col - mr.min_col + 1
        merges[f"{r0}_{c0}"] = {"r": r0, "c": c0, "rs": rs, "cs": cs}

    columnlen = {}
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        if cd and cd.width:
            columnlen[c - 1] = int(float(cd.width) * 7)

    sheet = {"name": ws.title, "data": data,
             "config": {"merge": merges, "columnlen": columnlen, "rowlen": {}}}
    return {"info": {"name": ws.title},
            "sheets": [sheet],
            "defaults": {"rowHeightPx": uniform_row_px, "colWidthPx": default_col_px}}

# --------- Key Notes helpers ----------
def _norm(s: str) -> str:
    return "".join(ch.lower() for ch in s if ch.isalnum())

def find_keynotes_row(ws, fallback=41):
    target = _norm("Key Notes/Follow Up")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if _norm(str(v)) == target:
                return r
    return fallback

def find_upstairs_row(ws):
    target = _norm("Upstairs")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if _norm(str(v)) == target:
                return r
    return None

def collect_existing_note_lines(ws, hdr_row, start_col=1, end_col=5, rows_below=4):
    lines = []
    for r in range(hdr_row + 1, hdr_row + rows_below + 1):
        parts = []
        for c in range(start_col, end_col + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                parts.append(str(v))
        lines.append(" ".join(parts).strip())
    while lines and lines[-1] == "":
        lines.pop()
    return "\n".join(lines).strip()

def shape_notes_exact(ws, hdr_row, start_col=1, end_col=5):
    lower = hdr_row + 1
    upper = hdr_row + 6
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= upper and mr.max_row >= lower:
            ws.unmerge_cells(range_string=mr.coord)

    ws.merge_cells(start_row=hdr_row + 1, start_column=start_col,
                   end_row=hdr_row + 4, end_column=end_col)
    ws.merge_cells(start_row=hdr_row + 5, start_column=start_col,
                   end_row=hdr_row + 5, end_column=end_col)
    ws.merge_cells(start_row=hdr_row + 6, start_column=start_col,
                   end_row=hdr_row + 6, end_column=end_col)

    return start_col, end_col, hdr_row + 1, hdr_row + 5, hdr_row + 6

def color_upstairs_black(ws, end_col=5):
    """Force all text in the 'Upstairs' section (until Key Notes header) to black."""
    up_hdr = find_upstairs_row(ws)
    if not up_hdr:
        return
    start_row = up_hdr + 1
    end_row = find_keynotes_row(ws, fallback=41) - 1
    if end_row < start_row:
        return
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            f = cell.font or Font()
            cell.font = Font(
                name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                underline=f.underline, strike=f.strike, vertAlign=f.vertAlign,
                color="000000"  # black
            )

# ---------------- Core ----------------
def apply_cells_and_export(cells, banner=None):
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME] if (SHEET_NAME and SHEET_NAME in wb.sheetnames) else wb.active

    original_merges = snapshot_merges(ws)
    unmerge_all(ws)

    # Write edits safely (to anchor if it was in a merged area)
    for cell in cells or []:
        r = int(cell.get("r", 0) or 0)
        c = int(cell.get("c", 0) or 0)
        v = cell.get("v", "")
        if r <= 0 or c <= 0:
            continue
        tr, tc = map_to_anchor_with_snapshot(r, c, original_merges)
        val = v
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                if not ALLOW_BLANK_OVERWRITE:
                    continue
                val = ""
            else:
                try:
                    if s.isdigit():
                        val = int(s)
                    else:
                        val = float(s)
                except Exception:
                    val = v
        ws.cell(row=tr, column=tc, value=val)

    # Banner text
    bn = banner or {}
    last_item     = (bn.get("lastNightItem") or "").strip()
    last_heavy    = (bn.get("lastNightHeavy") or "").strip()
    tonight_item  = (bn.get("tonightItem") or "").strip()
    tonight_heavy = (bn.get("tonightHeavy") or "").strip()

    top_lines = [
        "Good morning team.",
        f"Last night we took a {last_item}, it was heavy in {last_heavy}."
    ]
    bottom_lines = [
        f"Tonight we will be taking a {tonight_item}, it will be heavy in {tonight_heavy}.",
        "Please let me know if you have any questions. Thank you."
    ]

    top_count = len([t for t in top_lines if t.strip()])
    if top_count:
        ws.insert_rows(1, amount=top_count)
        _, used_c = _last_used_bounds(ws)
        end_c = max(used_c, 5)
        for i, text in enumerate(top_lines, start=1):
            if not text.strip():
                continue
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=end_c)
            ws.cell(i, 1, text).alignment = Alignment(vertical="center", wrap_text=True)

    shifted_merges = adjust_merges_row_offset(original_merges, top_count)
    reapply_merges(ws, shifted_merges)

    # Notes area (rows 42–45 merged, then rows 46–47 for sentences)
    hdr_row = find_keynotes_row(ws, fallback=41)
    collected = collect_existing_note_lines(ws, hdr_row, start_col=1, end_col=5, rows_below=4)

    start_c, end_c, notes_anchor_row, sent1_row, sent2_row = shape_notes_exact(ws, hdr_row, 1, 5)

    # --- NEW: write the textarea value if provided; else keep collected ---
    notes_text = (bn.get("notes") or bn.get("keyNotes") or bn.get("keynotes") or "").strip()
    if not notes_text:
        notes_text = collected

    anchor = ws.cell(notes_anchor_row, start_c)
    anchor.value = notes_text
    anchor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, text in ((sent1_row, bottom_lines[0]), (sent2_row, bottom_lines[1])):
        c = ws.cell(r, start_c)
        c.value = text
        c.alignment = Alignment(vertical="center", wrap_text=True)

    # Make the whole Upstairs block black
    color_upstairs_black(ws, end_col=5)

    # Autosize rows
    base_pt = _uniform_row_height_pt(ws)
    for rr in range(1, ws.max_row + 1):
        max_lines = 1
        for cc in range(1, ws.max_column + 1):
            cell = ws.cell(row=rr, column=cc)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None or str(val) == "":
                continue
            span = 1
            for mr in ws.merged_cells.ranges:
                if mr.min_row == rr and mr.min_col == cc:
                    span = mr.max_col - mr.min_col + 1
                    break
            chars_capacity = 0.0
            for ccc in range(cc, cc + span):
                chars_capacity += _col_char_width(ws, ccc)
            chars_capacity = max(1, int(round(chars_capacity)))
            needed = _estimate_needed_lines(str(val), chars_capacity)
            if needed > 1:
                cell.alignment = Alignment(wrap_text=True)
            max_lines = max(max_lines, needed)
        ws.row_dimensions[rr].height = max(base_pt, base_pt * max_lines)
    ws.sheet_format.defaultRowHeight = base_pt

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_filename = f"backstock-report-filled-{ts}.xlsx"
    return out, out_filename

# ---------------- Email ----------------
def send_via_smtp(file_bytes, filename, recipient, subject, body):
    if not recipient:
        raise RuntimeError("No recipient provided.")
    missing = [k for k, v in {
        "SENDER_EMAIL": SENDER_EMAIL,
        "SMTP_SERVER": SMTP_SERVER,
        "SMTP_USERNAME": SMTP_USERNAME,
        "SMTP_PASSWORD": SMTP_PASSWORD,
    }.items() if not v]
    if missing:
        raise RuntimeError(f"SMTP config missing: {', '.join(missing)}")

    msg = EmailMessage()
    msg["From"] = SENDER_EMAIL
    msg["To"] = recipient
    msg["Subject"] = subject or "Backstock Report"
    msg.set_content(body or "Please see the attached report.")
    msg.add_attachment(
        file_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(msg)

# ---------------- Routes ----------------
@app.route("/", methods=["GET"])
def home():
    return render_template("index.html", default_recipient=DEFAULT_RECIPIENT)

@app.route("/json", methods=["GET"])
def json_sheet():
    if not os.path.exists(TEMPLATE_PATH):
        return ("Template not found", 404)
    try:
        export_json = wb_to_luckysheet_json(TEMPLATE_PATH, SHEET_NAME)
        return app.response_class(
            response=json.dumps(export_json, default=str),
            status=200,
            mimetype="application/json",
            headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                     "Pragma": "no-cache", "Expires": "0"}
        )
    except Exception as e:
        return (f"Failed to convert workbook: {e}", 500)

@app.route("/download", methods=["POST"])
def download():
    payload = request.get_json(silent=True) or {}
    cells = payload.get("cells", [])
    banner = payload.get("banner", {})
    try:
        out, out_filename = apply_cells_and_export(cells, banner=banner)
        return send_file(out, as_attachment=True, download_name=out_filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return (f"Download failed: {e}", 500)

@app.route("/email", methods=["POST"])
def email():
    payload = request.get_json(silent=True) or {}
    cells = payload.get("cells", [])
    banner = payload.get("banner", {})
    recipient = (payload.get("recipient") or "").strip()
    subject = payload.get("subject") or "Backstock Report"
    body = payload.get("body") or "Please see the attached report."
    try:
        out, out_filename = apply_cells_and_export(cells, banner=banner)
        send_via_smtp(out.getvalue(), out_filename, recipient, subject, body)
        return jsonify({"ok": True, "message": f"Email sent to {recipient}."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/envcheck")
def envcheck():
    keys = ["SENDER_EMAIL", "SMTP_SERVER", "SMTP_PORT", "SMTP_USERNAME", "SMTP_PASSWORD"]
    return {k: bool(os.environ.get(k)) for k in keys}

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=True)
